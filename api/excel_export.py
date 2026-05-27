"""
Excel Export Module — generates .xlsx files compatible with ingestion templates.

Three exports:
- Prezzi:         asset_prices for assets in the portfolio
- Cedole e Fees:  dividends/expenses for the portfolio
- Transazioni:    buy/sell transactions for the portfolio
"""

import io
from datetime import datetime

import pandas as pd

try:
    from api.db_helper import execute_request
    from api.logger import logger
except ImportError:
    from db_helper import execute_request
    from logger import logger


def _fetch(table: str, params: dict) -> list:
    res = execute_request(table, 'GET', params=params)
    if res and res.status_code == 200:
        return res.json()
    return []


def _parse_date(d_str):
    if not d_str:
        return None
    try:
        # DB format is usually 'YYYY-MM-DD' or 'YYYY-MM-DDTHH:MM:SS...'
        return datetime.strptime(str(d_str).split('T')[0], '%Y-%m-%d').date()
    except Exception:
        return d_str


def _to_excel(df: pd.DataFrame, table_name: str = "TabellaDati", style_name: str = None, totals_row: bool = False) -> io.BytesIO:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        
        # Format date columns using Excel's built-in date format index 14 ('mm-dd-yy')
        # This maps directly to the standard Date category (responding to OS regional settings)
        date_headers = {'Data', 'Data Flusso', 'Data (acquisto/vendita)', 'Data Apertura', 'Data Chiusura'}
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in date_headers:
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = 'dd/mm/yyyy'
                    
        # Add Excel Table if not empty
        if not df.empty:
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
            
            last_col = get_column_letter(len(df.columns))
            # If totals row is shown, table range includes one extra row at the bottom
            extra_rows = 2 if totals_row else 1
            ref_range = f"A1:{last_col}{len(df) + extra_rows}"
            
            tab = Table(displayName=table_name, ref=ref_range)
            
            table_columns = []
            for col_idx, col_name in enumerate(df.columns, start=1):
                if totals_row and col_idx == 1:
                    # Prima colonna: etichetta "Totale" nella riga dei totali
                    tc = TableColumn(id=col_idx, name=col_name, totalsRowLabel='Totale')
                    worksheet.cell(row=len(df)+2, column=col_idx, value='Totale')
                elif totals_row and col_name in {'Dividendi', 'Controvalore'}:
                    # Colonne somma: totalsRowFunction imposta i metadati XML,
                    # ma occorre anche scrivere la formula nella cella affinché
                    # Excel mostri il valore senza richiedere F9.
                    # Con totalsRowCount=1, i riferimenti strutturati come [Dividendi]
                    # nelle celle della riga totali escludono la riga stessa,
                    # quindi SUBTOTAL NON genera circular reference.
                    tc = TableColumn(id=col_idx, name=col_name, totalsRowFunction='sum')
                    worksheet.cell(
                        row=len(df)+2, column=col_idx,
                        value=f'=SUBTOTAL(109,[{col_name}])'
                    )
                else:
                    tc = TableColumn(id=col_idx, name=col_name)
                table_columns.append(tc)
                
            tab.tableColumns = table_columns
            
            if totals_row:
                # totalsRowCount=1 è l'attributo OOXML che attiva la Total Row nativa
                # di Excel (spunta il checkbox "Total Row" nel tab Table Design).
                # totalsRowShown da solo NON produce questo effetto.
                tab.totalsRowCount = 1
            
            # Plain style with no colors (None) or standard Light1 style
            style = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True if style_name is not None else False,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
    buf.seek(0)
    return buf


def export_prezzi(portfolio_id: str) -> io.BytesIO:
    """
    Exports asset prices for all assets that appear in the given portfolio.
    Columns match PortfolioMP_Prezzi_Templates.xlsx.
    """
    # Fetch distinct ISINs in the portfolio via transactions
    tx = _fetch('transactions', {
        'portfolio_id': f'eq.{portfolio_id}',
        'select': 'assets(isin, name)'
    })

    # Build ISIN → name map
    isin_map: dict[str, str] = {}
    for t in tx:
        a = t.get('assets') or {}
        isin = a.get('isin')
        name = a.get('name', '')
        if isin:
            isin_map[isin] = name

    if not isin_map:
        df = pd.DataFrame(columns=['ISIN', 'Descrizione Asset', 'Data', 'Prezzo Corrente (EUR)'])
        return _to_excel(df, table_name="TabellaPrezzi")

    # Fetch prices for those ISINs
    # Supabase REST supports `in` filter with comma-separated values
    isin_list = ','.join(isin_map.keys())
    prices = _fetch('asset_prices', {
        'isin': f'in.({isin_list})',
        'select': 'isin,price,date',
        'order': 'isin,date'
    })

    rows = []
    for p in prices:
        isin = p.get('isin', '')
        date_val = _parse_date(p.get('date', ''))
        rows.append({
            'ISIN': isin,
            'Descrizione Asset': isin_map.get(isin, ''),
            'Data': date_val,
            'Prezzo Corrente (EUR)': p.get('price', '')
        })

    df = pd.DataFrame(rows, columns=['ISIN', 'Descrizione Asset', 'Data', 'Prezzo Corrente (EUR)'])
    return _to_excel(df, table_name="TabellaPrezzi")


def export_cedole(portfolio_id: str) -> io.BytesIO:
    """
    Exports dividends and fees for the portfolio.
    Columns match Portfolio_Cedole_Template.xlsx.
    """
    divs = _fetch('dividends', {
        'portfolio_id': f'eq.{portfolio_id}',
        'select': 'amount_eur,date,type,assets(isin,name)',
        'order': 'date'
    })

    rows = []
    for d in divs:
        a = d.get('assets') or {}
        amount = d.get('amount_eur')
        try:
            val = float(amount) if amount is not None else 0.0
        except (ValueError, TypeError):
            val = 0.0
        
        fees_val = 'Fee' if val < 0 else 'Cedole'
        date_val = _parse_date(d.get('date', ''))
        rows.append({
            'ISIN': a.get('isin', ''),
            'Valore Cedola (EUR)': amount,
            'Data Flusso': date_val,
            'Descrizione Titolo': a.get('name', ''),
            'Fees': fees_val
        })

    df = pd.DataFrame(rows, columns=['ISIN', 'Valore Cedola (EUR)', 'Data Flusso', 'Descrizione Titolo', 'Fees'])
    return _to_excel(df, table_name="TabellaCedole")


def export_transazioni(portfolio_id: str) -> io.BytesIO:
    """
    Exports buy/sell transactions for the portfolio.
    Columns match Portfolio_AcquistiVendite_Template.xlsx.
    """
    tx = _fetch('transactions', {
        'portfolio_id': f'eq.{portfolio_id}',
        'select': 'quantity,price_eur,date,type,assets(isin,name,asset_class)',
        'order': 'date'
    })

    rows = []
    for i, t in enumerate(tx):
        a = t.get('assets') or {}
        qty = t.get('quantity', 0)
        price = t.get('price_eur', 0)
        row_num = i + 2
        controvalore = f"=C{row_num}*E{row_num}"
        operazione = 'Acquisto' if t.get('type') == 'BUY' else 'Vendita'
        date_val = _parse_date(t.get('date', ''))
        rows.append({
            'ISIN': a.get('isin', ''),
            'Descrizione Asset': a.get('name', ''),
            'Quantità': qty,
            'Data (acquisto/vendita)': date_val,
            'Prezzo Operazione (EUR)': price,
            'Tipologia': a.get('asset_class', ''),
            'Operazione': operazione,
            'Divisa': 'EUR',
            'Controvalore (EUR)': controvalore,
            'Note': ''
        })

    cols = [
        'ISIN', 'Descrizione Asset', 'Quantità', 'Data (acquisto/vendita)',
        'Prezzo Operazione (EUR)', 'Tipologia', 'Operazione', 'Divisa',
        'Controvalore (EUR)', 'Note'
    ]
    df = pd.DataFrame(rows, columns=cols)
    return _to_excel(df, table_name="TabellaTransazioni")


def export_memory(portfolio_id: str) -> io.BytesIO:
    """
    Exports Note & Storico data to Excel.
    """
    try:
        from api.memory import compute_memory_data
    except ImportError:
        from memory import compute_memory_data
        
    data = compute_memory_data(portfolio_id)
    
    rows = []
    for d in data:
        # Format dates using _parse_date
        open_date = _parse_date(d.get('open_date'))
        close_date = _parse_date(d.get('close_date'))
        
        # P&L formatting as numeric float
        pnl = 0.0
        if d.get('pnl') is not None:
            try:
                pnl = float(d['pnl'])
            except (ValueError, TypeError):
                pass
                
        # Value formatting as numeric float
        value = 0.0
        if d.get('value') is not None:
            try:
                value = float(d['value'])
            except (ValueError, TypeError):
                pass
                
        # Total divs formatting as numeric float
        total_divs = 0.0
        if d.get('total_divs') is not None:
            try:
                total_divs = float(d['total_divs'])
            except (ValueError, TypeError):
                pass

        # MWR formatted string (simple/period/annual type)
        mwr_val = ''
        mwr = d.get('mwr')
        mwr_type = d.get('mwr_type', 'NONE')
        if mwr_type not in ['NONE', 'ERROR'] and mwr is not None:
            suffix = ""
            if mwr_type == "SIMPLE": suffix = " (S)"
            elif mwr_type == "PERIOD": suffix = " (P)"
            elif mwr_type == "ANNUAL": suffix = " (A)"
            try:
                mwr_val = f"{float(mwr):.2f}%{suffix}"
            except (ValueError, TypeError):
                mwr_val = str(mwr)
        
        rows.append({
            'Descrizione Asset': d.get('description', ''),
            'ISIN': d.get('isin', ''),
            'Tipologia': d.get('type', ''),
            'P&L': pnl,
            'MWR%': mwr_val,
            'Dividendi': total_divs,
            'Controvalore': value,
            'Data Apertura': open_date,
            'Data Chiusura': close_date,
            'Note': d.get('note', '')
        })
        
    cols = [
        'Descrizione Asset', 'ISIN', 'Tipologia', 'P&L', 'MWR%', 
        'Dividendi', 'Controvalore', 'Data Apertura', 'Data Chiusura', 'Note'
    ]
    df = pd.DataFrame(rows, columns=cols)
    
    # Simple table with style 'TableStyleLight1' as requested, and totals row enabled
    return _to_excel(df, table_name="TabellaStorico", style_name="TableStyleLight1", totals_row=True)
